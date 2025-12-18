import os
import asyncio
import tempfile
from datetime import datetime, time

from aiogram import Bot, Dispatcher, types
from aiogram.filters import CommandStart
from aiogram.types import Message
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties

from openpyxl import load_workbook
from openpyxl.styles import Font

# ================= CONFIG =================

BOT_TOKEN = os.getenv("BOT_TOKEN")

WHITE_FONT = Font(color="FFFFFF")

# ================= SAFE HELPERS =================

def safe_str(v):
    return "" if v is None else str(v)

def set_cell(ws, r, c, value):
    ws.cell(r, c).value = value
    ws.cell(r, c).font = WHITE_FONT

def get_last_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            return r
    return 1

def excel_time_to_hms(v):
    if v is None or v == "":
        return ""

    if isinstance(v, datetime):
        return v.strftime("%H:%M:%S")

    if isinstance(v, time):
        return v.strftime("%H:%M:%S")

    if isinstance(v, (int, float)):
        seconds = int(round(v * 86400))
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    try:
        return datetime.strptime(str(v)[:8], "%H:%M:%S").strftime("%H:%M:%S")
    except:
        return ""

# ================= EXCEL PROCESS =================

def process_excel_file(input_path, output_path):
    wb = load_workbook(input_path, data_only=True)
    ws = wb["Sheet1"]

    if "constant" in wb.sheetnames:
        wb.remove(wb["constant"])

    constant = wb.create_sheet("constant")

    last_row = get_last_row(ws)

    for i in range(1, last_row + 1):
        set_cell(constant, i, 119, safe_str(ws.cell(i, 19).value))
        set_cell(constant, i, 120, safe_str(ws.cell(i, 20).value))
        set_cell(constant, i, 121, safe_str(ws.cell(i, 21).value))

        ora = excel_time_to_hms(ws.cell(i, 2).value)
        set_cell(constant, i, 122, ora)

        set_cell(constant, i, 123, safe_str(ws.cell(i, 23).value))

        text_note = (
            safe_str(ws.cell(i, 21).value)
            or safe_str(ws.cell(i, 20).value)
            or safe_str(ws.cell(i, 19).value)
        )

        if ora and text_note:
            set_cell(constant, i, 124, f"{ora} {text_note}")

        set_cell(
            constant,
            i,
            125,
            f"{safe_str(ws.cell(i, 19).value)} {safe_str(ws.cell(i, 21).value)}".strip(),
        )

    wb.save(output_path)

# ================= TELEGRAM BOT =================

bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML),
)

dp = Dispatcher()

@dp.message(CommandStart())
async def start(msg: Message):
    await msg.answer("👋 Trimite fișier Excel (.xls / .xlsx / .xlsm)")

@dp.message(lambda m: m.document is not None)
async def handle_excel(msg: Message):
    doc = msg.document
    name = doc.file_name.lower()

    if not name.endswith((".xls", ".xlsx", ".xlsm")):
        await msg.answer("❌ Format invalid.")
        return

    await msg.answer("📥 Am primit fișierul. Procesez...")

    with tempfile.TemporaryDirectory() as tmp:
        inp = os.path.join(tmp, doc.file_name)
        out = os.path.join(tmp, "modificat_" + doc.file_name)

        await bot.download(doc, destination=inp)
        process_excel_file(inp, out)

        await msg.answer_document(
            types.FSInputFile(out),
            caption="✅ Gata. Fișier procesat.",
        )

# ================= RUN =================

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
