import os
from datetime import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xlrd  # pentru .xls

# ================== STILURI ==================

YELLOW = "FFFF00"
GREEN_DARK = "00B050"   # RGB(0,176,80)
GREEN_LIME = "00FF00"   # RGB(0,255,0)
RED = "FF0000"

font12_bold = Font(name="Arial", size=12, bold=True)
font14_bold = Font(name="Arial", size=14, bold=True)

fill_yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
fill_green = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
fill_red = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
fill_lime = PatternFill(start_color=GREEN_LIME, end_color=GREEN_LIME, fill_type="solid")

align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
no_border = Border()


# ================== HELPERI ==================

def nz(v):
    if v is None:
        return ""
    try:
        s = str(v)
    except Exception:
        return ""
    return s.strip()


def last_data_row(ws, col_idx: int) -> int:
    last = ws.max_row
    while last > 0:
        v = ws.cell(row=last, column=col_idx).value
        if v not in (None, "") and str(v).strip() != "":
            return last
        last -= 1
    return 1


def convert_xls_to_xlsx(xls_path: str) -> str:
    """
    Conversie minimala .xls -> .xlsx (valori doar).
    Folosita DOAR pentru fisiere primite ca .xls.
    Originalul .xls NU este modificat.
    """
    if not os.path.isfile(xls_path):
        raise ValueError("Fisier .xls inexistent.")
    base, _ = os.path.splitext(xls_path)
    new_path = base + "_conv.xlsx"

    book = xlrd.open_workbook(xls_path, formatting_info=False)
    wb = Workbook()
    # stergem sheet-ul gol creat automat
    wb.remove(wb.active)

    for si in range(book.nsheets):
        sh = book.sheet_by_index(si)
        title = (sh.name or f"Sheet{si+1}")[:31]
        ws = wb.create_sheet(title=title)
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                ws.cell(row=r + 1, column=c + 1).value = sh.cell_value(r, c)

    wb.save(new_path)
    return new_path


# ==========================================================
# 1) PUB_Zero -> PUB_IN  (ConvertAndFormatCols7_8_9_10_11)
#    DOAR coloanele G-K, restul ramane intact
# ==========================================================

def format_pub_zero(pub_zero_path: str) -> str:
    if not os.path.isfile(pub_zero_path):
        raise ValueError("Fisierul PUB_Zero nu exista.")

    original_path = pub_zero_path
    base_dir = os.path.dirname(original_path)
    base_name, ext = os.path.splitext(os.path.basename(original_path))

    # daca e .xls -> lucram pe copie convertita
    if ext.lower() == ".xls":
        pub_zero_path = convert_xls_to_xlsx(original_path)
        base_name, ext = os.path.splitext(os.path.basename(pub_zero_path))

    if ext.lower() not in (".xlsx", ".xlsm"):
        raise ValueError("PUB_Zero trebuie sa fie .xlsx / .xlsm / .xls")

    keep_vba = (ext.lower() == ".xlsm")
    wb = load_workbook(pub_zero_path, keep_vba=keep_vba)
    if not wb.worksheets:
        wb.close()
        raise ValueError("PUB_Zero nu contine foi.")
    ws = wb.worksheets[0]

    last_row = last_data_row(ws, 7)
    if last_row <= 1:
        wb.close()
        raise ValueError("Nu sunt date in coloana G pentru PUB_Zero.")

    for row in range(2, last_row + 1):
        c7 = ws.cell(row=row, column=7)
        c8 = ws.cell(row=row, column=8)
        c9 = ws.cell(row=row, column=9)
        c10 = ws.cell(row=row, column=10)
        c11 = ws.cell(row=row, column=11)

        # ---- G: stil + galben ----
        c7.font = font12_bold
        c7.fill = fill_yellow
        c7.alignment = align_left

        # ---- H: doar daca numeric >0 -> HH:MM:SS, apoi stil ----
        val_h = c8.value
        if isinstance(val_h, (int, float)) and val_h > 0:
            total = int(val_h)
            h = total // 3600
            m = (total % 3600) // 60
            s = total % 60
            c8.value = f"{h:02d}:{m:02d}:{s:02d}"
        c8.font = font12_bold
        c8.fill = fill_yellow
        c8.alignment = align_right

        # ---- I: doar daca nu e gol ----
        if nz(c9.value) != "":
            c9.font = font14_bold
            c9.fill = fill_green
            c9.alignment = align_center
            c9.border = thin_border
        else:
            c9.border = no_border

        # ---- J: prefix/sufix "_" ----
        if nz(c10.value) != "":
            s_val = nz(c10.value)
            if s_val.isdigit() and 1 <= int(s_val) <= 9:
                c10.value = f"_{s_val}__"
            else:
                c10.value = f"_{s_val}_"
        c10.font = font14_bold
        c10.fill = fill_yellow
        c10.alignment = align_center

        # ---- K: border doar daca are continut ----
        if nz(c11.value) != "":
            c11.border = thin_border
        else:
            c11.border = no_border

    # Nume ca in macro: _modificat + aceeasi extensie (xls -> xlsx)
    out_ext = ".xlsx" if ext.lower() == ".xls" else ext or ".xlsx"
    out_name = f"{base_name}_modificat{out_ext}"
    out_path = os.path.join(base_dir, out_name)
    wb.save(out_path)
    wb.close()

    return out_path  # acesta este PUB_IN


# ==========================================================
# 2) FLOW COMBINAT (IN + PUB_IN) -> IN_modificat
# ==========================================================

def col4_excluded(val: str) -> bool:
    v = nz(val).lower()

    if v.startswith("id pub") or v.startswith("id_pub_") \
       or v.startswith("id promo") or v.startswith("id_promo_") \
       or v.startswith("interzis") or v.startswith("cca_") \
       or v.startswith("cca orele"):
        return True

    excl = {
        "id_jtv_2024_dua_lipa_dance_the_night",
        "id_jtv_2024_miley_cyrus_flowers",
        "id_jtv_2024_the weeknd_ariana grande_save_your_tears",
        "id 15 ani_25sec_v1",
        "youtube sofia obada jurnalul orei 19 ok",
        "jurnalsportiv",
        "meteonew",
    }
    return v in excl


def color_red_col_e(ws):
    last = max(last_data_row(ws, 4), last_data_row(ws, 5))
    for r in range(1, last + 1):
        val_d = nz(ws.cell(r, 4).value)
        val_e = nz(ws.cell(r, 5).value)
        if val_e != "" and not col4_excluded(val_d):
            c = ws.cell(r, 5)
            c.fill = fill_red  # doar culoarea, ca in macro


def delete_between_playlist_markers(ws):
    last = last_data_row(ws, 6)
    i = 1
    while i <= last:
        val = nz(ws.cell(i, 6).value)
        if val.startswith("PLAYLIST_IN_"):
            start_row = i
            end_row = 0
            j = i + 1
            while j <= last:
                v2 = nz(ws.cell(j, 6).value)
                if v2.startswith("PLAYLIST_OUT_"):
                    end_row = j
                    break
                j += 1
            if end_row > 0:
                if end_row > start_row + 1:
                    ws.delete_rows(start_row + 1, end_row - start_row - 1)
                    last = last_data_row(ws, 6)
                i = start_row
            else:
                break
        i += 1


def intervals_def():
    raw = [
        "06:00:00,06:30:00,06_30,06_20,06_10", "06:30:01,06:59:00,06_50,06_40,06_45",
        "07:00:00,07:30:00,07_20,07_10,07_30", "07:31:00,07:59:00,07_50,07_40,07_45",
        "08:00:00,08:31:00,08_20,08_10,08_30", "08:32:00,08:59:00,08_50,08_40,08_45",
        "09:00:00,09:31:00,09_20,09_10,09_30", "09:32:00,09:59:00,09_50,09_40,09_45",
        "10:00:00,10:31:00,10_20,10_10,10_30", "10:32:00,10:59:00,10_50,10_40,10_45",
        "11:00:00,11:31:00,11_20,11_10,11_30", "11:32:00,11:59:00,11_50,11_40,11_45",
        "12:00:00,12:31:00,12_20,12_10,12_30", "12:32:00,12:59:00,12_50,12_40,12_45",
        "13:00:00,13:31:00,13_20,13_10,13_30", "13:32:00,13:59:00,13_50,13_40,13_45",
        "14:00:00,14:31:00,14_20,14_10,14_30", "14:32:00,14:59:00,14_50,14_40,14_45",
        "15:00:00,15:31:00,15_20,15_10,15_30", "15:32:00,15:59:00,15_50,15_40,15_45",
        "16:00:00,16:31:00,16_20,16_10,16_30", "16:32:00,16:59:00,16_50,16_40,16_45",
        "17:00:00,17:31:00,17_20,17_10,17_30", "17:32:00,17:59:00,17_50,17_40,17_45",
        "18:00:00,18:31:00,18_20,18_10,18_30", "18:32:00,18:59:00,18_50,18_40,18_45",
        "19:00:00,19:31:00,19_20,19_10,19_30", "19:32:00,19:59:00,19_50,19_40,19_45",
        "20:00:00,20:31:00,20_20,20_10,20_30", "20:32:00,20:59:00,20_50,20_40,20_45",
        "21:00:00,21:31:00,21_20,21_10,21_30", "21:32:00,21:59:00,21_50,21_40,21_45",
        "22:00:00,22:31:00,22_20,22_10,22_30", "22:32:00,22:59:00,22_50,22_40,22_45",
        "23:00:00,23:31:00,23_20,23_10,23_30", "23:32:00,23:59:00,23_50,23_40,23_45",
        "00:00:00,00:31:00,00_20,00_10,00_30", "00:32:00,00:59:00,00_50,00_40,00_45",
        "01:00:00,01:31:00,01_20,01_10,01_30", "01:32:00,01:59:00,01_50,01_40,01_45",
    ]
    out = []
    for row in raw:
        parts = row.split(",")
        out.append((parts[0], parts[1], parts[2:]))
    return out


def time_in_range(t, start_str, end_str):
    def to_time(s):
        h, m, s2 = map(int, s.split(":"))
        return time(h, m, s2)
    start = to_time(start_str)
    end = to_time(end_str)
    return start <= t <= end


def collect_block_for_interval(ws_in, ora_start, ora_end):
    last_in = last_data_row(ws_in, 3)
    iduri, colG, colH, colI = [], [], [], []

    r = 1
    while r <= last_in:
        txt = nz(ws_in.cell(r, 3).value)
        if txt and ":" in txt:
            try:
                h, m, s = map(int, txt.split(":"))
                t = time(h, m, s)
            except Exception:
                t = None
            if t and time_in_range(t, ora_start, ora_end):
                rr = r
                while rr <= last_in:
                    txt2 = nz(ws_in.cell(rr, 3).value)
                    if txt2 and ":" in txt2:
                        try:
                            hh, mm, ss = map(int, txt2.split(":"))
                            tt = time(hh, mm, ss)
                        except Exception:
                            tt = None
                        if tt and not time_in_range(tt, ora_start, ora_end):
                            break
                    if nz(ws_in.cell(rr, 10).value) != "":
                        iduri.append(nz(ws_in.cell(rr, 10).value))
                        colG.append(nz(ws_in.cell(rr, 7).value))
                        colH.append(nz(ws_in.cell(rr, 8).value))
                        colI.append(nz(ws_in.cell(rr, 9).value))
                    rr += 1
                break
        r += 1

    return iduri, colG, colH, colI


def apply_block_to_out(ws_out, iduri, colG, colH, colI, variante):
    if not iduri:
        return

    last_out = last_data_row(ws_out, 6)

    for var in variante:
        tag_in = f"PLAYLIST_IN_{var}"
        tag_out = f"PLAYLIST_OUT_{var}"
        start_row = 0
        end_row = 0

        for r in range(1, last_out + 1):
            val = nz(ws_out.cell(r, 6).value)
            if val == tag_in:
                start_row = r
            if val == tag_out and start_row > 0:
                end_row = r
                break

        if start_row > 0 and end_row > start_row:
            if end_row > start_row + 1:
                ws_out.delete_rows(start_row + 1, end_row - start_row - 1)
                end_row = start_row + 1

            ws_out.insert_rows(end_row, amount=len(iduri))

            for i in range(len(iduri)):
                r_ins = start_row + 1 + i

                # D = G, E = H, F = ID, G = I (stiluri ca in macro)
                c4 = ws_out.cell(r_ins, 4)
                c4.value = colG[i]
                c4.font = font14_bold
                c4.fill = fill_yellow
                c4.alignment = align_left

                c5 = ws_out.cell(r_ins, 5)
                c5.value = colH[i]
                c5.font = font14_bold
                c5.fill = fill_yellow
                c5.alignment = align_right

                c6 = ws_out.cell(r_ins, 6)
                c6.value = iduri[i]
                c6.font = font14_bold
                c6.fill = fill_yellow
                c6.alignment = align_center

                c7 = ws_out.cell(r_ins, 7)
                c7.value = colI[i]
                if nz(colI[i]) != "":
                    c7.fill = fill_lime
                    c7.border = thin_border

            break  # doar prima varianta valida


def process_all_intervals(ws_in, ws_out):
    for ora_start, ora_end, variante in intervals_def():
        iduri, colG, colH, colI = collect_block_for_interval(ws_in, ora_start, ora_end)
        if iduri:
            apply_block_to_out(ws_out, iduri, colG, colH, colI, variante)


def run_combined_flow(in_path: str, pub_in_path: str) -> str:
    """
    Proceseaza_Flow_Combinat + SaveWithSuffixKeepExt:
    - nu atinge fisierul original IN
    - returneaza copie <IN>_modificat.(ext sau .xlsx)
    """
    if not os.path.isfile(in_path):
        raise ValueError("Fisierul IN nu exista.")
    if not os.path.isfile(pub_in_path):
        raise ValueError("Fisierul PUB_IN nu exista.")

    original_in = in_path
    base_dir = os.path.dirname(original_in)
    base_name, ext = os.path.splitext(os.path.basename(original_in))

    # .xls -> convertim copie
    if ext.lower() == ".xls":
        in_path = convert_xls_to_xlsx(original_in)
        out_ext = ".xlsx"
    else:
        out_ext = ext or ".xlsx"

    keep_vba = (ext.lower() == ".xlsm")
    wb_out = load_workbook(in_path, keep_vba=keep_vba)
    if not wb_out.worksheets:
        wb_out.close()
        raise ValueError("IN nu contine foi.")
    ws_out = wb_out.worksheets[0]

    # PAS 1: colorare rosu pe E
    color_red_col_e(ws_out)

    # PAS 2: sterge intre PLAYLIST_IN_/PLAYLIST_OUT_
    delete_between_playlist_markers(ws_out)

    # PAS 2b: PUB_IN
    wb_in = load_workbook(pub_in_path)
    if not wb_in.worksheets:
        wb_in.close()
        wb_out.close()
        raise ValueError("PUB_IN nu contine foi.")
    ws_in = wb_in.worksheets[0]

    # PAS 2c: intervale
    process_all_intervals(ws_in, ws_out)

    wb_in.close()

    # PAS 3: salvare ca in SaveWithSuffixKeepExt
    final_name = f"{base_name}_modificat{out_ext}"
    final_path = os.path.join(base_dir, final_name)
    wb_out.save(final_path)
    wb_out.close()

    return final_path
